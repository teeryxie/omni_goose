#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import shutil
from datetime import time
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any
from zipfile import ZipFile

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT_DIR = ROOT / "data_extend"
DEFAULT_OUTPUT_DIR = ROOT / "data" / "level_2_extended"

PACKAGE_SPECS = {
    "level_2_luo_23.zip": {
        "prefix": "luo23",
        "language_group": "mixed_luo_23",
    },
    "yang.zip": {
        "prefix": "yang_gen",
        "language_group": "mixed_yang_gen",
    },
}

HEADER_ALIASES = {
    "video_id": "video_id",
    "quesetion_1": "question_1",
    "question_time": "question_time",
    "A": "option_A",
    "B": "option_B",
    "正确答案": "correct_answer",
    "quesetion_2": "question_2",
    "current_ASR": "current_asr",
    "涉及人数": "num_people",
    "交替性": "turn_taking",
    "光照": "lighting",
    "一致性": "consistency",
    "相机情况": "camera",
    "ASR": "asr",
    "备注": "language_note",
}


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\u2003", " ").strip()


def _format_timestamp(value: Any) -> str:
    if isinstance(value, time):
        return f"{value.hour:02d}:{value.minute:02d}:{value.second:02d}"
    if isinstance(value, (int, float)):
        if float(value).is_integer():
            return str(int(value))
        return str(value)
    return _clean_text(value)


def _header_map(headers: list[Any]) -> dict[str, int]:
    mapped: dict[str, int] = {}
    for idx, raw in enumerate(headers):
        name = HEADER_ALIASES.get(_clean_text(raw))
        if name and name not in mapped:
            mapped[name] = idx
    return mapped


def _cell(row: tuple[Any, ...], indexes: dict[str, int], name: str) -> Any:
    idx = indexes.get(name)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _video_id(value: Any) -> int | None:
    if value is None:
        return None
    text = _clean_text(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _build_sample(row: tuple[Any, ...], indexes: dict[str, int], prefix: str, language_group: str) -> dict[str, Any] | None:
    source_video_id = _video_id(_cell(row, indexes, "video_id"))
    if source_video_id is None:
        return None

    correct_answer = _clean_text(_cell(row, indexes, "correct_answer")).upper()
    if correct_answer not in {"A", "B"}:
        raise ValueError(f"{prefix}/{source_video_id}: invalid correct_answer={correct_answer!r}")

    q1 = {
        "question": _clean_text(_cell(row, indexes, "question_1")),
        "timestamp": _format_timestamp(_cell(row, indexes, "question_time")),
        "option_A": _clean_text(_cell(row, indexes, "option_A")) or "YES",
        "option_B": _clean_text(_cell(row, indexes, "option_B")) or "NO",
        "correct_answer": correct_answer,
    }
    q2 = {
        "question": _clean_text(_cell(row, indexes, "question_2")),
        "answer": _clean_text(_cell(row, indexes, "current_asr")),
    }

    return {
        "video_id": f"{prefix}_{source_video_id:03d}",
        "video_file": f"{prefix}_{source_video_id:03d}.mp4",
        "source_package": prefix,
        "source_video_id": source_video_id,
        "language_group": language_group,
        "language_note": _clean_text(_cell(row, indexes, "language_note")),
        "question_1": q1,
        "question_2": q2,
        "metadata": {
            "num_people": _cell(row, indexes, "num_people"),
            "turn_taking": _clean_text(_cell(row, indexes, "turn_taking")),
            "lighting": _clean_text(_cell(row, indexes, "lighting")),
            "consistency": _clean_text(_cell(row, indexes, "consistency")),
            "camera": _clean_text(_cell(row, indexes, "camera")),
            "asr": _clean_text(_cell(row, indexes, "asr")),
        },
    }


def _read_package(zip_path: Path, spec: dict[str, str], videos_dir: Path) -> list[dict[str, Any]]:
    samples: list[dict[str, Any]] = []
    with TemporaryDirectory() as tmp:
        tmp_dir = Path(tmp)
        with ZipFile(zip_path) as archive:
            names = archive.namelist()
            xlsx_names = [name for name in names if name.lower().endswith(".xlsx")]
            if len(xlsx_names) != 1:
                raise ValueError(f"{zip_path}: expected exactly one xlsx, got {xlsx_names}")
            xlsx_path = tmp_dir / xlsx_names[0]
            archive.extract(xlsx_names[0], tmp_dir)

            sheet = load_workbook(xlsx_path, data_only=True).active
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            indexes = _header_map(headers)

            for row in sheet.iter_rows(min_row=2, values_only=True):
                sample = _build_sample(row, indexes, spec["prefix"], spec["language_group"])
                if sample is None:
                    continue
                source_video = next(
                    (name for name in names if name.endswith(f"/{sample['source_video_id']}.mp4")),
                    None,
                )
                if source_video is None:
                    raise FileNotFoundError(f"{zip_path}: missing video {sample['source_video_id']}.mp4")
                target_video = videos_dir / sample["video_file"]
                with archive.open(source_video) as src, target_video.open("wb") as dst:
                    shutil.copyfileobj(src, dst)
                samples.append(sample)
    return samples


def prepare(input_dir: Path, output_dir: Path) -> Path:
    videos_dir = output_dir / "videos"
    videos_dir.mkdir(parents=True, exist_ok=True)

    samples: list[dict[str, Any]] = []
    for zip_name, spec in PACKAGE_SPECS.items():
        zip_path = input_dir / zip_name
        if not zip_path.exists():
            raise FileNotFoundError(f"Missing source zip: {zip_path}")
        samples.extend(_read_package(zip_path, spec, videos_dir))

    payload = {
        "dataset": "SocialOmni Level 2 Extended",
        "version": "2026-05-06",
        "source_dir": str(input_dir),
        "total": len(samples),
        "data": samples,
    }
    output_path = output_dir / "annotations.json"
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return output_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Prepare extended Level 2 multilingual data")
    parser.add_argument("--input-dir", type=Path, default=DEFAULT_INPUT_DIR)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output_path = prepare(args.input_dir, args.output_dir)
    print(output_path)


if __name__ == "__main__":
    main()
